import re
import time
from datetime import datetime, timedelta
from threading import Thread

import telebot
from telebot import types
from openpyxl import load_workbook

#Bot token
bot = telebot.TeleBot('6482401483:AAGfjUM6u6BJuCz0KJ1V5fLcEr_o1yLKaLI')

worker_dict = {'userid': {'userid_saving': None, #айди пользователя
                          'username': None, #ФИО пользователя
                          'starttime': None, #Время начала рабочего дня
                          'endtime': None, #Время окончания рабочего дня
                          'overtime': None #Переработка
                            }}


def overtime_check(time, userid):
    try:
        if time.hour >= 19 or time.hour <= 5:
            try:
                worker_dict[userid]['overtime'] = True
            except KeyError:
                worker_dict.update({userid:{'overtime': True}})
        if worker_dict[userid]['overtime'] is None:
            worker_dict[userid]['overtime'] = False
    except UnboundLocalError:
        print('Не удалось зафиксировать переработку: UnboundLocalError')
    

def write_in_excel(userid):
    wb = load_workbook('WorkTimeControl.xlsx')
    ws = wb[worker_dict[userid]['starttime'].strftime('%B')]
    print('текущий лист: ', ws.title)
    #номер строки, в которой находится userid в эксель-таблице
    row_number = None
    for rows in ws['A']:
        if rows.value == userid:
            row_number = rows.row
            index2 = 'B' + str(row_number)
            ws[index2] = worker_dict[userid]['username']
            break
        elif not rows.value:
            row_number = rows.row
            index = 'A' + str(row_number)
            index2 = 'B' + str(row_number)
            ws[index] = userid
            ws[index2] = worker_dict[userid]['username']
            break
    if row_number is None:
        row_number = rows.row + 1
    #Столбец, в первой строке которого написано текущее число (17, например)
    for row in ws['1']:
        if row.value == int(worker_dict[userid]['starttime'].day):
            column_letter_ = row.column_letter
            break
    #создаём координату ячейки по номеру строки с userid и столбцу с текущей датой
    coord_cell = column_letter_ + str(row_number)
    print('Координата определена:', coord_cell)
    if ws[coord_cell].value is not None:#если в ячейке уже что-то есть, вытаскиваем оттуда время, которое пользователь провёл на работе, складываем с новым значением из calc()
        if worker_dict[userid]['overtime'] == True:
            current_value = ws[coord_cell].value
            # добавить к разнице между началом и концом рабочего дня часы которые отработаны.  
            current_value += worker_dict[userid]['endtime'] - worker_dict[userid]['starttime']
        elif (worker_dict[userid]['overtime'] == False and worker_dict[userid]['endtime']-worker_dict[userid]['starttime'] >= timedelta(hours=8))\
            or (ws[coord_cell].value >= timedelta(hours=8) and worker_dict[userid]['overtime'] == False):
            current_value = worker_dict[userid]['endtime'] - worker_dict[userid]['starttime']
            #Если текущее значение больше 8:30, вычитаем час в качестве времени на обед.  
            if current_value > timedelta(hours=8,minutes=30):
                current_value -= timedelta(hours=1)
                print('Обед просчитан')
        elif (worker_dict[userid]['overtime'] == False and worker_dict[userid]['endtime']-worker_dict[userid]['starttime'] < timedelta(hours=8))\
            or (ws[coord_cell].value < timedelta(hours=8) and worker_dict[userid]['overtime'] == False):
            current_value = ws[coord_cell].value
            # добавить к разнице между началом и концом рабочего дня часы которые отработаны.
            current_value += worker_dict[userid]['endtime'] - worker_dict[userid]['starttime']
        ws[coord_cell] = current_value #записываем полученное количество рабочего времени в эксель
        print('Данные сохранены:', worker_dict[userid]['username'], '. Провёл на работе:', current_value, '. Ячейка:', coord_cell)
    else:
        data_to_save = worker_dict[userid]['endtime'] - worker_dict[userid]['starttime']
        if data_to_save > timedelta(hours=8,minutes=30):
                data_to_save -= timedelta(hours=1)
                print('Обед просчитан')
        ws[coord_cell] = data_to_save
        #если в таблице значения нет - записываем разницу конец рабочего дня - начало рабочего дня
        print('Данные сохранены:', worker_dict[userid]['username'], '. Провёл на работе:', data_to_save, '. Ячейка:', coord_cell)
    wb.save('WorkTimeControl.xlsx')


def usr_btns(message):
    #Отдельная клава для юзеров и админов. В соновном, для выгрузки в лс
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton("+о")
    btn2 = types.KeyboardButton("-о")
    btn3 = types.KeyboardButton("+д")
    btn4 = types.KeyboardButton("-д")
    if message.from_user.id in [400585102, 321166597]:
        btn6 = types.KeyboardButton('Выгрузка')
        markup.add(btn1, btn2, btn3, btn4, btn6)
    else:
        markup.add(btn1, btn2, btn3, btn4)
    return markup


@bot.message_handler(commands=['start'])
def start(message):
    markup = usr_btns(message)
    dialog_id = message.chat.id
    bot.send_message(dialog_id, 'Кронос', reply_markup=markup) 

@bot.message_handler(content_types=['text'])
def get_text_messages(message):
    # Сохраняем id диалога в dialg_id.
    dialog_id = message.chat.id
    # Получаем ФИО пользователя для записи в эксель.
    if message.from_user.first_name is not None and message.from_user.last_name is not None:
        user = message.from_user.first_name + ' ' + message.from_user.last_name
    elif message.from_user.first_name is not None:
        user = message.from_user.first_name
    elif message.from_user.last_name is not None:
        user = message.from_user.last_name
    else: user = message.from_user.first_name + ' ' + message.from_user.last_name 
    # Получаем уникальный идентификатор пользователя из телеграма в userid.
    userid = message.from_user.id
    try:
        worker_dict[userid]['username'] = user
    except KeyError:
        worker_dict.update({userid:{'userid_saving': userid,
                                    'username': user,
                                    'overtime': None,
                                    'starttime': None,
                                    'endtime': None}})
    # В зависимости от айди пользователя, создаём разную клавиатуру.  
    msg = message.text.lower() # Приводим текст сообщения к нижнему регистру.
    msg = msg.replace(' ', '') # Убираем все пробелы из сообщения.
    msg = msg.replace('^', ':')
    msg = msg.replace('+0', '+о')
    msg_str_dat = re.findall(r"([0-1][0-9]|2[0-3]|[0-9])\S([0-5][0-9]|[0-9])", str(msg))
    try:
        if not re.findall(r"(\+|-|_|=)(д|l|о|j)\D*([0-1][0-9]|2[0-3]|[0-9])\S([0-5][0-9]|[0-9])", str(msg)):
            msg = re.findall(r"(\+|-|_|=)(д|l|о|j)", str(msg))
            if message.text.startswith('+')  or message.text.startswith('='):
                try:
                    worker_dict[userid]['starttime'] = datetime.now()
                    worker_dict[userid]['endtime'] = None
                except KeyError:
                    worker_dict.update({userid:{'userid_saving': userid,
                                                'username': user, 
                                                'starttime' : datetime.now(),
                                                'endtime': None
                                                }})
                overtime_check(datetime.now(), userid)
                print(f'{user} - {message.text} - {datetime.now()}')
                # Если регулярка высосала 1 символ после -, сохраняем datetime в словарь.  
            elif message.text.startswith('-')  or message.text.startswith('_'):
                worker_dict[userid]['endtime'] = datetime.now()
                if worker_dict[userid]['overtime'] is not None and worker_dict[userid]['starttime'] is not None:
                    write_in_excel(userid)
                    worker_dict[userid]['starttime'] = None
                elif worker_dict[userid]['starttime'] is None:
                    print(f'{user} - не найдено время начала рабочего дня')
                print(f'{user} - {message.text} - {datetime.now()}')
        else:
            msg_new_data = [i for j in msg_str_dat for i in j]
            print(f'{user} - {message.text} - Время: {msg_new_data}')
            if len(msg_new_data) > 0:
                    # если регулярка высосала время из сообщения, приводим его к формату datetime.  
                datetime_from_msg = datetime(year=datetime.now().year, \
                                             month=datetime.now().month,\
                                             day=datetime.now().day,\
                                             hour=int(msg_new_data[0]),\
                                             minute=int(msg_new_data[1]), second=0)
            msg = re.findall(r"(\+|-|_|=)(д|l|о|j)\D*([0-1][0-9]|2[0-3]|[0-9])\S([0-5][0-9]|[0-9]\b)", str(msg))
            if len(msg) > 0:
                if message.text.startswith('+')  or message.text.startswith('='):
                    try:
                        worker_dict[userid]['starttime'] = datetime_from_msg
                    except KeyError:
                        worker_dict.update({userid:{'userid_saving': userid,
                                                    'username': user, 
                                                    'starttime' : datetime_from_msg,
                                                    'endtime': None
                                                    }})
                    overtime_check(datetime_from_msg, userid)
                elif message.text.startswith('-')  or message.text.startswith('_'):
                    worker_dict[userid]['endtime'] = datetime_from_msg
                    if worker_dict[userid]['overtime'] is not None and worker_dict[userid]['starttime'] is not None:
                        write_in_excel(userid)
                        worker_dict[userid]['starttime'] = None
                    elif worker_dict[userid]['starttime'] is None:
                        print(f'{user} - не найдено время начала рабочего дня')
    except IndexError:
        print(f'Сообщение не обработано: {user}, {message.text} -> {msg}')
    if message.text.lower() == 'выгрузка':
        bot.send_document(dialog_id, open(r'WorkTimeControl.xlsx', 'rb'))


def daily_worker():
    global worker_dict
    while True:
        now = datetime.now()
        time_to_sleep = (datetime(year=now.year, month=now.month, day=now.day+1, hour=4, minute=30, second=30) - now).seconds
        print(f'[Daily Worker] Sleeping for {time_to_sleep} seconds')
        time.sleep(time_to_sleep)
        now = datetime.now()
        print(f'[Daily Worker] Saving all who had worked for 8 hours...')
        for key in worker_dict.keys():
            if worker_dict[key]['endtime'] is None and worker_dict[key]['starttime'] is not None:
                worker_dict[key]['endtime'] = worker_dict[key]['starttime'] + timedelta(hours=8)
                write_in_excel(key)
        print(f'[Daily Worker] Saved all who had worked for 8 hours.')
        print(f'[Daily Worker] Now clearing worker_dict.')
        worker_dict.clear()


daily_thread = Thread(target=daily_worker)
daily_thread.start()
bot.infinity_polling(timeout=None)   
